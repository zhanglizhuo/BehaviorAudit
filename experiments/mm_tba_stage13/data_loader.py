import os
import re
from pathlib import Path

import numpy as np
import openpyxl

FEATURE_NAMES = [
    'char_count',
    'sentence_count',
    'avg_sentence_length',
    'question_count',
    'exclamation_count',
    'pause_marker_count',
    'digit_count',
    'math_keyword_count',
    'gender',
    'qualification',
    'experience_years',
    'grade_index',
    'subject_index',
]

MATH_KEYWORDS = ['函数', '方程', '不等式', '角', '圆', '分数', '概率', '几何', '证明']
GRADE_MAP = {
    'p1': 1,
    'p2': 2,
    'p3': 3,
    'p4': 4,
    'p5': 5,
    'p6': 6,
    'j1': 7,
    'j2': 8,
    'j3': 9,
    's1': 10,
    's2': 11,
    's3': 12,
}
SUBJECT_MAP = {
    'math': 1,
    'mathematics': 1,
    'english': 2,
    'chinese': 3,
    'physics': 4,
    'chemistry': 5,
    'biology': 6,
    'history': 7,
    'geography': 8,
    'politics': 9,
}

def _candidate_dataset_roots(root_dir='.'):
    env_root = os.environ.get('MM_TBA_ROOT')
    if env_root:
        yield Path(env_root).expanduser().resolve()

    start = Path(root_dir).resolve()
    for base in [start, *start.parents]:
        yield base / 'MM-TBA'

    current_file = Path(__file__).resolve()
    for base in current_file.parents:
        yield base / 'MM-TBA'

def _find_dataset_root(root_dir='.'):
    checked = []
    for candidate in _candidate_dataset_roots(root_dir):
        if candidate in checked:
            continue
        checked.append(candidate)
        if (candidate / 'Teacher_Lecture_Evaluation').exists():
            return candidate
    return None

def _parse_experience_years(value):
    if value is None:
        return 0.0
    text = str(value).strip().lower()
    if not text:
        return 0.0
    match = re.search(r'(\d+(?:\.\d+)?)\s*([a-z]+)', text)
    if not match:
        try:
            return float(text)
        except ValueError:
            return 0.0
    amount = float(match.group(1))
    unit = match.group(2)
    if unit.startswith('m'):
        return amount / 12.0
    return amount

def _parse_score_token(token):
    cleaned = token.replace(' ', '')
    if '~' in cleaned:
        left, right = cleaned.split('~', 1)
        return (float(left) + float(right)) / 2.0
    if '-' in cleaned:
        left, right = cleaned.split('-', 1)
        return (float(left) + float(right)) / 2.0
    return float(cleaned)

def _extract_scores(report_text):
    raw_tokens = re.findall(r'分数[:：]\s*([0-9]+(?:\.[0-9]+)?(?:\s*[~-]\s*[0-9]+(?:\.[0-9]+)?)?)', report_text)
    scores = []
    for token in raw_tokens:
        try:
            scores.append(_parse_score_token(token))
        except ValueError:
            continue
    return scores

def _split_sentences(text):
    return [segment for segment in re.split(r'[。！？!?]+', text) if segment.strip()]

def _load_metadata(metadata_path):
    workbook = openpyxl.load_workbook(metadata_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return {}

    header = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
    metadata = {}
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        record = dict(zip(header, row))
        sample_id = str(record.get('Filename')).strip()
        if not sample_id:
            continue
        metadata[sample_id] = record
    return metadata

def _subject_index(value):
    text = str(value or '').strip().lower()
    return SUBJECT_MAP.get(text, 0)

def _grade_index(value):
    text = str(value or '').strip().lower()
    return GRADE_MAP.get(text, 0)

def _text_features(transcript, metadata):
    transcript = transcript.strip()
    sentences = _split_sentences(transcript)
    char_count = len(transcript)
    sentence_count = len(sentences)
    avg_sentence_length = float(char_count / sentence_count) if sentence_count else 0.0
    question_count = transcript.count('？') + transcript.count('?')
    exclamation_count = transcript.count('！') + transcript.count('!')
    pause_marker_count = transcript.count('嗯') + transcript.count('啊') + transcript.count('呃')
    digit_count = sum(ch.isdigit() for ch in transcript)
    math_keyword_count = sum(transcript.count(keyword) for keyword in MATH_KEYWORDS)

    gender = float(metadata.get('Gender') or 0)
    qualification = float(metadata.get('Teacher Qualification Certificate') or 0)
    experience_years = _parse_experience_years(metadata.get('Teaching Experience'))
    grade_index = float(_grade_index(metadata.get('Grade')))
    subject_index = float(_subject_index(metadata.get('Subject')))

    return [
        float(char_count),
        float(sentence_count),
        avg_sentence_length,
        float(question_count),
        float(exclamation_count),
        float(pause_marker_count),
        float(digit_count),
        float(math_keyword_count),
        gender,
        qualification,
        experience_years,
        grade_index,
        subject_index,
    ]

def _build_samples(dataset_root):
    lecture_root = dataset_root / 'Teacher_Lecture_Evaluation'
    transcript_dir = lecture_root / 'teacher_lecture_texts'
    metadata_path = dataset_root / 'metadata.xlsx'
    report_dirs = [lecture_root / 'gpt_report' / 'train', lecture_root / 'gpt_report' / 'eval']

    metadata_map = _load_metadata(metadata_path)
    samples = []
    found_files = [str(metadata_path)]

    for report_dir in report_dirs:
        if not report_dir.exists():
            continue
        found_files.append(str(report_dir))
        for report_path in sorted(report_dir.glob('*.txt')):
            sample_id = report_path.stem
            transcript_path = transcript_dir / f'{sample_id}.txt'
            if not transcript_path.exists():
                continue

            report_text = report_path.read_text(encoding='utf-8', errors='ignore')
            transcript_text = transcript_path.read_text(encoding='utf-8', errors='ignore')
            scores = _extract_scores(report_text)
            if len(scores) < 4:
                continue

            metadata = metadata_map.get(sample_id, {})
            samples.append(
                {
                    'lecture_id': sample_id,
                    'target': float(np.mean(scores[:4])),
                    'features': _text_features(transcript_text, metadata),
                    'metadata': metadata,
                }
            )

    return samples, found_files

def load_mm_tba_data(root_dir='.'):
    dataset_root = _find_dataset_root(root_dir)
    data_status = {
        'dataset_root': str(dataset_root) if dataset_root else None,
        'found_files': [],
        'total_samples': 0,
        'features': FEATURE_NAMES,
        'target': 'mean_rubric_score',
        'teacher_ids': None,
        'missing_data': False,
        'samples': [],
        'synthetic_fallback': False,
        'error': None,
    }

    if dataset_root is None:
        data_status['missing_data'] = True
        data_status['error'] = 'MM-TBA dataset root not found. Set MM_TBA_ROOT or place MM-TBA inside the repository root.'
        return data_status

    samples, found_files = _build_samples(dataset_root)
    data_status['found_files'] = found_files
    data_status['samples'] = samples
    data_status['total_samples'] = len(samples)

    if not samples:
        data_status['missing_data'] = True
        data_status['error'] = 'No matched MM-TBA lecture evaluation samples were found.'

    return data_status

def get_data_card(data_status):
    return {
        'dataset_name': 'MM-TBA Teacher Lecture Evaluation',
        'dataset_root': data_status.get('dataset_root'),
        'found_files': data_status.get('found_files', []),
        'total_samples': data_status.get('total_samples', 0),
        'missing_data': data_status.get('missing_data', True),
        'synthetic_fallback': False,
        'target_name': data_status.get('target'),
        'num_features': len(data_status.get('features', [])),
        'feature_names': data_status.get('features', []),
        'has_teacher_ids': False,
        'num_teachers': 0,
        'annotation_note': 'Samples are built from MM-TBA lecture transcripts, GPT rubric reports, and metadata.xlsx. No synthetic fallback is allowed.',
        'error': data_status.get('error'),
    }

def extract_features_and_targets(data_status, root_dir='.'):
    del root_dir
    samples = data_status.get('samples') or []
    if data_status.get('missing_data') or not samples:
        return None, None, None

    X = np.array([sample['features'] for sample in samples], dtype=float)
    y = np.array([sample['target'] for sample in samples], dtype=float)
    return X, y, None