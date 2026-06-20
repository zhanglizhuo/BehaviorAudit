from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import numpy as np
import openpyxl

from ..types import AuditDatasetBundle


FEATURE_NAMES = [
    "char_count",
    "sentence_count",
    "avg_sentence_length",
    "question_count",
    "exclamation_count",
    "pause_marker_count",
    "digit_count",
    "math_keyword_count",
    "gender",
    "qualification",
    "experience_years",
    "grade_index",
    "subject_index",
]

MATH_KEYWORDS = ["函数", "方程", "不等式", "角", "圆", "分数", "概率", "几何", "证明"]
GRADE_MAP = {
    "p1": 1,
    "p2": 2,
    "p3": 3,
    "p4": 4,
    "p5": 5,
    "p6": 6,
    "j1": 7,
    "j2": 8,
    "j3": 9,
    "s1": 10,
    "s2": 11,
    "s3": 12,
}
SUBJECT_MAP = {
    "math": 1,
    "mathematics": 1,
    "english": 2,
    "chinese": 3,
    "physics": 4,
    "chemistry": 5,
    "biology": 6,
    "history": 7,
    "geography": 8,
    "politics": 9,
}


class MMTBAAdapter:
    name = "mm_tba"

    def load(self, dataset_root: Optional[str] = None) -> AuditDatasetBundle:
        resolved_root = self._find_dataset_root(dataset_root)
        if resolved_root is None:
            error = "MM-TBA dataset root not found. Set MM_TBA_ROOT or provide --dataset-root."
            return AuditDatasetBundle(
                dataset_name="MM-TBA Teacher Lecture Evaluation",
                dataset_root=None,
                X=None,
                y=None,
                group_ids=None,
                data_card=self._data_card(None, [], 0, True, error),
                feature_names=FEATURE_NAMES,
                missing_data=True,
                error=error,
            )

        samples, found_files = self._build_samples(resolved_root)
        if not samples:
            error = "No matched MM-TBA lecture evaluation samples were found."
            return AuditDatasetBundle(
                dataset_name="MM-TBA Teacher Lecture Evaluation",
                dataset_root=str(resolved_root),
                X=None,
                y=None,
                group_ids=None,
                data_card=self._data_card(str(resolved_root), found_files, 0, True, error),
                feature_names=FEATURE_NAMES,
                missing_data=True,
                error=error,
            )

        X = np.asarray([sample["features"] for sample in samples], dtype=float)
        y = np.asarray([sample["target"] for sample in samples], dtype=float)
        data_card = self._data_card(str(resolved_root), found_files, len(samples), False, None)
        return AuditDatasetBundle(
            dataset_name="MM-TBA Teacher Lecture Evaluation",
            dataset_root=str(resolved_root),
            X=X,
            y=y,
            group_ids=None,
            data_card=data_card,
            feature_names=FEATURE_NAMES,
            missing_data=False,
            error=None,
        )

    def _candidate_dataset_roots(self, dataset_root: Optional[str]) -> List[Path]:
        candidates: List[Path] = []
        if dataset_root:
            candidates.append(Path(dataset_root).expanduser().resolve())

        env_root = os.environ.get("MM_TBA_ROOT")
        if env_root:
            candidates.append(Path(env_root).expanduser().resolve())

        behavioraudit_root = Path(__file__).resolve().parents[2]
        repo_root = behavioraudit_root.parent
        for base in [Path.cwd().resolve(), behavioraudit_root, repo_root, *list(repo_root.parents)[:2]]:
            candidates.append(base / "MM-TBA")

        unique_candidates: List[Path] = []
        seen: Set[Path] = set()
        for candidate in candidates:
            if candidate not in seen:
                unique_candidates.append(candidate)
                seen.add(candidate)
        return unique_candidates

    def _find_dataset_root(self, dataset_root: Optional[str]) -> Optional[Path]:
        for candidate in self._candidate_dataset_roots(dataset_root):
            if (candidate / "Teacher_Lecture_Evaluation").exists():
                return candidate
        return None

    def _build_samples(self, dataset_root: Path) -> Tuple[List[Dict[str, object]], List[str]]:
        lecture_root = dataset_root / "Teacher_Lecture_Evaluation"
        transcript_dir = lecture_root / "teacher_lecture_texts"
        metadata_path = dataset_root / "metadata.xlsx"
        report_dirs = [lecture_root / "gpt_report" / "train", lecture_root / "gpt_report" / "eval"]

        metadata_map = self._load_metadata(metadata_path)
        samples: List[Dict[str, object]] = []
        found_files = [str(metadata_path)]

        for report_dir in report_dirs:
            if not report_dir.exists():
                continue
            found_files.append(str(report_dir))
            for report_path in sorted(report_dir.glob("*.txt")):
                sample_id = report_path.stem
                transcript_path = transcript_dir / f"{sample_id}.txt"
                if not transcript_path.exists():
                    continue

                report_text = report_path.read_text(encoding="utf-8", errors="ignore")
                transcript_text = transcript_path.read_text(encoding="utf-8", errors="ignore")
                scores = self._extract_scores(report_text)
                if len(scores) < 4:
                    continue

                metadata = metadata_map.get(sample_id, {})
                samples.append(
                    {
                        "lecture_id": sample_id,
                        "target": float(np.mean(scores[:4])),
                        "features": self._text_features(transcript_text, metadata),
                    }
                )

        return samples, found_files

    def _load_metadata(self, metadata_path: Path) -> Dict[str, Dict[str, object]]:
        workbook = openpyxl.load_workbook(metadata_path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return {}

        header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        metadata: Dict[str, Dict[str, object]] = {}
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            record = dict(zip(header, row))
            sample_id = str(record.get("Filename") or "").strip()
            if sample_id:
                metadata[sample_id] = record
        return metadata

    def _parse_experience_years(self, value: object) -> float:
        if value is None:
            return 0.0
        text = str(value).strip().lower()
        if not text:
            return 0.0
        match = re.search(r"(\d+(?:\.\d+)?)\s*([a-z]+)", text)
        if not match:
            try:
                return float(text)
            except ValueError:
                return 0.0
        amount = float(match.group(1))
        unit = match.group(2)
        return amount / 12.0 if unit.startswith("m") else amount

    def _parse_score_token(self, token: str) -> float:
        cleaned = token.replace(" ", "")
        if "~" in cleaned:
            left, right = cleaned.split("~", 1)
            return (float(left) + float(right)) / 2.0
        if "-" in cleaned:
            left, right = cleaned.split("-", 1)
            return (float(left) + float(right)) / 2.0
        return float(cleaned)

    def _extract_scores(self, report_text: str) -> List[float]:
        raw_tokens = re.findall(r"分数[:：]\s*([0-9]+(?:\.[0-9]+)?(?:\s*[~-]\s*[0-9]+(?:\.[0-9]+)?)?)", report_text)
        scores: List[float] = []
        for token in raw_tokens:
            try:
                scores.append(self._parse_score_token(token))
            except ValueError:
                continue
        return scores

    def _split_sentences(self, text: str) -> List[str]:
        return [segment for segment in re.split(r"[。！？!?]+", text) if segment.strip()]

    def _subject_index(self, value: object) -> float:
        return float(SUBJECT_MAP.get(str(value or "").strip().lower(), 0))

    def _grade_index(self, value: object) -> float:
        return float(GRADE_MAP.get(str(value or "").strip().lower(), 0))

    def _text_features(self, transcript: str, metadata: Dict[str, object]) -> List[float]:
        transcript = transcript.strip()
        sentences = self._split_sentences(transcript)
        char_count = len(transcript)
        sentence_count = len(sentences)
        avg_sentence_length = float(char_count / sentence_count) if sentence_count else 0.0
        question_count = transcript.count("？") + transcript.count("?")
        exclamation_count = transcript.count("！") + transcript.count("!")
        pause_marker_count = transcript.count("嗯") + transcript.count("啊") + transcript.count("呃")
        digit_count = sum(ch.isdigit() for ch in transcript)
        math_keyword_count = sum(transcript.count(keyword) for keyword in MATH_KEYWORDS)

        return [
            float(char_count),
            float(sentence_count),
            avg_sentence_length,
            float(question_count),
            float(exclamation_count),
            float(pause_marker_count),
            float(digit_count),
            float(math_keyword_count),
            float(metadata.get("Gender") or 0),
            float(metadata.get("Teacher Qualification Certificate") or 0),
            self._parse_experience_years(metadata.get("Teaching Experience")),
            self._grade_index(metadata.get("Grade")),
            self._subject_index(metadata.get("Subject")),
        ]

    def _data_card(
        self,
        dataset_root: Optional[str],
        found_files: List[str],
        total_samples: int,
        missing_data: bool,
        error: Optional[str],
    ) -> Dict[str, object]:
        return {
            "dataset_name": "MM-TBA Teacher Lecture Evaluation",
            "dataset_root": dataset_root,
            "found_files": found_files,
            "total_samples": total_samples,
            "missing_data": missing_data,
            "synthetic_fallback": False,
            "target_name": "mean_rubric_score",
            "num_features": len(FEATURE_NAMES),
            "feature_names": FEATURE_NAMES,
            "has_group_ids": False,
            "num_groups": 0,
            "group_type": None,
            "annotation_note": "Samples are built from MM-TBA lecture transcripts, GPT rubric reports, and metadata.xlsx. No synthetic fallback is allowed.",
            "error": error,
        }
